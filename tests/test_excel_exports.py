from datetime import datetime, timedelta
from decimal import Decimal
from io import BytesIO

import pytest
from django.contrib.auth import get_user_model
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook

from apps.fleet.models import EcoStandard, Transport
from apps.locations.models import Location
from apps.orders.models import OrderPoint, ShipmentOrder
from apps.routing.services.route_calculation_service import RouteCalculationService
from apps.trips.models import Trip
from apps.trips.services import TripLifecycleService

User = get_user_model()

EXPECTED_HEADERS = [
    "Дата рейса",
    "Менеджер",
    "Транспорт",
    "Евро-класс",
    "Маршрут",
    "Расстояние, км",
    "Время, мин",
    "Топливо, л",
    "Стоимость, руб",
    "CO2, кг",
    "NOx, г",
    "PM, г",
    "Эко-рейтинг",
    "CO2/км",
    "CO2/тонно-км",
    "Версия расчетной модели",
]
XLSX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


@pytest.fixture
def manager():
    return User.objects.create_user(
        username="excel_manager",
        password="StrongPass12345",
        role=User.Role.MANAGER,
        first_name="Иван",
        last_name="Менеджеров",
    )


@pytest.fixture
def other_manager():
    return User.objects.create_user(
        username="excel_other_manager",
        password="StrongPass12345",
        role=User.Role.MANAGER,
    )


@pytest.fixture
def admin_user():
    return User.objects.create_user(
        username="excel_admin",
        password="StrongPass12345",
        role=User.Role.ADMIN,
    )


@pytest.fixture
def transport():
    standard = EcoStandard.objects.create(
        name="Euro VI",
        nox_limit_g_per_kwh=Decimal("0.46"),
        pm_limit_mg_per_kwh=Decimal("10.00"),
    )
    return Transport.objects.create(
        plate_number="Е160ХХ777",
        model="КАМАЗ 5490",
        category=Transport.Category.N3,
        capacity_kg=20000,
        fuel_consumption_l_per_100km=Decimal("29.00"),
        eco_standard=standard,
        year=2021,
    )


@pytest.fixture
def locations():
    origin = Location.objects.create(
        name="Москва",
        latitude=Decimal("55.7558"),
        longitude=Decimal("37.6173"),
    )
    destination = Location.objects.create(
        name="Подольск",
        latitude=Decimal("55.4312"),
        longitude=Decimal("37.5447"),
    )
    return origin, destination


def create_trip(manager, transport, locations, route_name):
    origin, destination = locations
    order = ShipmentOrder.objects.create(
        manager=manager,
        transport=transport,
        cargo_name=f"Груз {route_name}",
        cargo_type="Паллеты",
        cargo_weight_kg=Decimal("1000.00"),
        desired_delivery_date="2026-06-01",
    )
    OrderPoint.objects.create(
        order=order,
        location=origin,
        sequence=1,
        point_type=OrderPoint.PointType.PICKUP,
    )
    OrderPoint.objects.create(
        order=order,
        location=destination,
        sequence=2,
        point_type=OrderPoint.PointType.DELIVERY,
    )
    RouteCalculationService().calculate_for_order(order)
    route = order.route_options.first()
    route.name = route_name
    route.save(update_fields=["name", "updated_at"])
    return TripLifecycleService().approve_route(order, route, manager)


def deliver_trip(trip, manager, finish_at):
    service = TripLifecycleService()
    service.start_trip(trip, manager, actual_start_at=finish_at - timedelta(hours=2))
    service.deliver_trip(trip, manager, actual_finish_at=finish_at)
    trip.refresh_from_db()
    return trip


def aware(year, month, day, hour=12):
    return timezone.make_aware(datetime(year, month, day, hour, 0))


def set_route_snapshot(
    trip,
    *,
    distance_km=Decimal("42.50"),
    duration_minutes=75,
    fuel_liters=Decimal("15.25"),
    cost_rub=Decimal("12345.67"),
    co2_kg=Decimal("40.12"),
    nox_g=Decimal("5.34"),
    pm_g=Decimal("0.123"),
    eco_rating=Decimal("88.80"),
    calculation_model_version="v2.1",
    calculation_details_json=None,
):
    route = trip.route_option
    route.distance_km = distance_km
    route.duration_minutes = duration_minutes
    route.fuel_liters = fuel_liters
    route.cost_rub = cost_rub
    route.co2_kg = co2_kg
    route.nox_g = nox_g
    route.pm_g = pm_g
    route.eco_rating = eco_rating
    route.calculation_model_version = calculation_model_version
    route.calculation_details_json = (
        {
            "co2_kg_per_km": "0.944",
            "co2_kg_per_ton_km": "0.1888",
            "calculation_model_version": calculation_model_version,
        }
        if calculation_details_json is None
        else calculation_details_json
    )
    route.save()


def load_response_workbook(response):
    assert response.status_code == 200
    assert response["Content-Type"] == XLSX_CONTENT_TYPE
    assert response.content.startswith(b"PK")
    return load_workbook(BytesIO(response.content), data_only=True)


def table_header(sheet):
    for row in sheet.iter_rows(values_only=True):
        if row and row[0] == "Дата рейса":
            return list(row[: len(EXPECTED_HEADERS)])
    raise AssertionError("Не найдена строка заголовков Excel-таблицы.")


def row_by_route(sheet, route_name):
    for row in sheet.iter_rows(values_only=True):
        if row and route_name in row:
            return list(row[: len(EXPECTED_HEADERS)])
    raise AssertionError(f"Не найдена строка маршрута {route_name}.")


def assert_route_absent(sheet, route_name):
    assert all(route_name not in row for row in sheet.iter_rows(values_only=True) if row)


@pytest.mark.django_db
def test_manager_emissions_xlsx_uses_saved_snapshots_scope_and_date_filters(
    client, manager, other_manager, transport, locations
):
    included = deliver_trip(
        create_trip(manager, transport, locations, "Маршрут Excel"),
        manager,
        aware(2026, 6, 10),
    )
    set_route_snapshot(included, co2_kg=Decimal("123.45"))
    old_snapshot = deliver_trip(
        create_trip(manager, transport, locations, "Старый снимок"),
        manager,
        aware(2026, 6, 15),
    )
    set_route_snapshot(old_snapshot, calculation_model_version="v1", calculation_details_json={})
    outside_period = deliver_trip(
        create_trip(manager, transport, locations, "Вне периода"),
        manager,
        aware(2026, 7, 1),
    )
    set_route_snapshot(outside_period)
    other_trip = deliver_trip(
        create_trip(other_manager, transport, locations, "Чужой маршрут"),
        other_manager,
        aware(2026, 6, 12),
    )
    set_route_snapshot(other_trip)
    client.force_login(manager)

    response = client.get(
        reverse("reports:emissions_xlsx"),
        {"date_from": "2026-06-01", "date_to": "2026-06-30"},
    )
    workbook = load_response_workbook(response)
    sheet = workbook["Отчет по выбросам"]

    assert table_header(sheet) == EXPECTED_HEADERS
    row = row_by_route(sheet, "Маршрут Excel")
    old_row = row_by_route(sheet, "Старый снимок")
    assert row[EXPECTED_HEADERS.index("Менеджер")] == "Иван Менеджеров"
    assert row[EXPECTED_HEADERS.index("CO2, кг")] == 123.45
    assert row[EXPECTED_HEADERS.index("CO2/км")] == 0.944
    assert row[EXPECTED_HEADERS.index("Версия расчетной модели")] == "v2.1"
    assert old_row[EXPECTED_HEADERS.index("CO2/км")] == "—"
    assert old_row[EXPECTED_HEADERS.index("CO2/тонно-км")] == "—"
    assert old_row[EXPECTED_HEADERS.index("Версия расчетной модели")] == "v1"
    assert_route_absent(sheet, "Вне периода")
    assert_route_absent(sheet, "Чужой маршрут")


@pytest.mark.django_db
def test_admin_company_xlsx_access_rules_and_company_scope(
    client, admin_user, manager, other_manager, transport, locations
):
    first = deliver_trip(
        create_trip(manager, transport, locations, "Компания 1"),
        manager,
        aware(2026, 6, 1),
    )
    set_route_snapshot(first)
    second = deliver_trip(
        create_trip(other_manager, transport, locations, "Компания 2"),
        other_manager,
        aware(2026, 6, 2),
    )
    set_route_snapshot(second)

    assert client.get(reverse("dashboard:admin_dashboard_xlsx")).status_code == 302
    client.force_login(manager)
    assert client.get(reverse("dashboard:admin_dashboard_xlsx")).status_code == 403
    client.force_login(admin_user)

    response = client.get(reverse("dashboard:admin_dashboard_xlsx"))
    workbook = load_response_workbook(response)
    sheet = workbook["Доставленные рейсы"]

    assert table_header(sheet) == EXPECTED_HEADERS
    assert row_by_route(sheet, "Компания 1")
    assert row_by_route(sheet, "Компания 2")


@pytest.mark.django_db
def test_trips_xlsx_respects_manager_scope_and_status_filter(
    client, manager, other_manager, transport, locations
):
    own_delivered = deliver_trip(
        create_trip(manager, transport, locations, "Свой доставлен"),
        manager,
        aware(2026, 6, 1),
    )
    set_route_snapshot(own_delivered)
    own_planned = create_trip(manager, transport, locations, "Свой план")
    set_route_snapshot(own_planned)
    other_delivered = deliver_trip(
        create_trip(other_manager, transport, locations, "Чужой доставлен"),
        other_manager,
        aware(2026, 6, 1),
    )
    set_route_snapshot(other_delivered)
    client.force_login(manager)

    response = client.get(reverse("trips:export_xlsx"), {"status": Trip.Status.DELIVERED})
    workbook = load_response_workbook(response)
    sheet = workbook["Рейсы"]

    assert table_header(sheet) == EXPECTED_HEADERS
    assert row_by_route(sheet, "Свой доставлен")
    assert_route_absent(sheet, "Свой план")
    assert_route_absent(sheet, "Чужой доставлен")
