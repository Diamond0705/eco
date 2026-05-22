from decimal import Decimal
from io import BytesIO

from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from apps.routing.services.route_snapshot_metrics import (
    average_decimal,
    calculation_model_version,
    co2_kg_per_km,
    co2_kg_per_ton_km,
    display_decimal,
    has_tolls,
)

MISSING_VALUE = "—"
XLSX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

TRIP_EXPORT_HEADERS = [
    "Дата рейса",
    "Менеджер",
    "Транспорт",
    "Евро-класс",
    "Маршрут",
    "Расстояние, км",
    "Время, мин",
    "Топливо, л",
    "Стоимость, руб",
    "CO2, кг",
    "NOx, г",
    "PM, г",
    "Эко-рейтинг",
    "CO2/км",
    "CO2/тонно-км",
    "Версия расчетной модели",
]

NUMERIC_FORMATS = {
    "Расстояние, км": "0.00",
    "Время, мин": "0",
    "Топливо, л": "0.00",
    "Стоимость, руб": "#,##0.00",
    "CO2, кг": "0.00",
    "NOx, г": "0.00",
    "PM, г": "0.000",
    "Эко-рейтинг": "0.00",
    "CO2/км": "0.000",
    "CO2/тонно-км": "0.0000",
}


def build_xlsx_response(xlsx_bytes, filename):
    response = HttpResponse(xlsx_bytes, content_type=XLSX_CONTENT_TYPE)
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


class TripExcelExportService:
    def build_emissions_report(self, manager, report):
        manager_name = _user_display_name(manager)
        filters = report["filters"]
        workbook = self._new_workbook("Отчет по выбросам")
        sheet = workbook.active
        sheet.append(["Отчет по выбросам"])
        sheet.append(["Менеджер", manager_name])
        sheet.append(["Период", self._period_label(filters)])
        sheet.append(["Сформировано", timezone.localtime().strftime("%d.%m.%Y %H:%M")])
        sheet.append([])
        self._write_summary(sheet, report["summary"])
        sheet.append([])
        self._write_trip_table(sheet, [row["trip"] for row in report["rows"]])
        return self._save(workbook)

    def build_company_dashboard(self, analytics):
        workbook = self._new_workbook("Сводка компании")
        summary_sheet = workbook.active
        summary_sheet.append(["Сводка компании"])
        summary_sheet.append(["Сформировано", timezone.localtime().strftime("%d.%m.%Y %H:%M")])
        summary_sheet.append([])
        summary_sheet.append(["Показатель", "Значение"])
        for label, value in self._company_summary_rows(analytics):
            summary_sheet.append([label, value])
        self._style_summary(summary_sheet)

        trips_sheet = workbook.create_sheet("Доставленные рейсы")
        trips = analytics.get("delivered_trips", [])
        self._write_trip_table(trips_sheet, trips)
        return self._save(workbook)

    def build_trips_export(self, trips):
        workbook = self._new_workbook("Рейсы")
        sheet = workbook.active
        sheet.append(["Экспорт рейсов"])
        sheet.append(["Сформировано", timezone.localtime().strftime("%d.%m.%Y %H:%M")])
        sheet.append([])
        self._write_trip_table(sheet, trips)
        return self._save(workbook)

    def _new_workbook(self, title):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = title
        return workbook

    def _write_summary(self, sheet, summary):
        sheet.append(["Итоговые показатели", ""])
        summary_rows = [
            ("Рейсы", summary["trips_count"]),
            ("Расстояние, км", summary["distance_km"]),
            ("Топливо, л", summary["fuel_liters"]),
            ("Стоимость, руб", summary["cost_rub"]),
            ("CO2, кг", summary["co2_kg"]),
            ("NOx, г", summary["nox_g"]),
            ("PM, г", summary["pm_g"]),
            ("Средний CO2/км", summary["average_co2_kg_per_km"]),
            ("Средний CO2/тонно-км", summary["average_co2_kg_per_ton_km"]),
            ("Средний эко-рейтинг", summary["average_eco_rating"]),
            ("Маршруты с платными участками", summary["toll_routes_count"]),
        ]
        for label, value in summary_rows:
            sheet.append([label, _cell_value(value)])
        self._style_summary(sheet)

    def _write_trip_table(self, sheet, trips):
        header_row = sheet.max_row + 1
        sheet.append(TRIP_EXPORT_HEADERS)
        for trip in trips:
            sheet.append(self._trip_row(trip))
        self._style_table(sheet, header_row)
        self._write_table_totals(sheet, header_row)

    def _trip_row(self, trip):
        order = trip.order
        transport = order.transport
        route = trip.route_option
        return [
            _excel_datetime(_trip_date(trip)),
            _user_display_name(order.manager),
            str(transport),
            transport.eco_standard.name,
            route.name,
            _number(route.distance_km),
            route.duration_minutes,
            _number(route.fuel_liters),
            _number(route.cost_rub),
            _number(route.co2_kg),
            _number(route.nox_g),
            _number(route.pm_g),
            _number(route.eco_rating),
            _optional_number(co2_kg_per_km(route)),
            _optional_number(co2_kg_per_ton_km(route)),
            calculation_model_version(route),
        ]

    def _write_table_totals(self, sheet, header_row):
        data_start = header_row + 1
        data_end = sheet.max_row
        if data_end < data_start:
            return

        totals = _summary_for_trips(sheet.iter_rows(min_row=data_start, max_row=data_end))
        sheet.append([])
        sheet.append(["Итого / среднее", "", "", "", ""])
        summary_row = sheet.max_row
        for column_index, header in enumerate(TRIP_EXPORT_HEADERS, start=1):
            value = totals.get(header)
            if value is None:
                continue
            cell = sheet.cell(summary_row, column_index, value)
            cell.number_format = NUMERIC_FORMATS.get(header, "General")
        sheet.cell(summary_row, 1).font = Font(bold=True)

    def _company_summary_rows(self, analytics):
        company = analytics["company"]
        return [
            ("Пользователи", analytics["users"]["total"]),
            ("Менеджеры", analytics["users"]["managers"]),
            ("Транспорт", analytics["transports"]["total"]),
            ("Заявки", analytics["orders"]["total"]),
            ("Рейсы", analytics["trips"]["total"]),
            ("Доставленные рейсы", analytics["trips"]["delivered"]),
            ("Расстояние, км", company["distance_km"]),
            ("Топливо, л", company["fuel_liters"]),
            ("Стоимость, руб", company["cost_rub"]),
            ("CO2, кг", company["co2_kg"]),
            ("NOx, г", company["nox_g"]),
            ("PM, г", company["pm_g"]),
            ("Средний CO2/км", company["average_co2_kg_per_km"]),
            ("Средний CO2/тонно-км", company["average_co2_kg_per_ton_km"]),
            ("Средний эко-рейтинг", company["average_eco_rating"]),
            ("Маршруты с платными участками", company["toll_routes_count"]),
        ]

    def _style_summary(self, sheet):
        sheet.column_dimensions["A"].width = 34
        sheet.column_dimensions["B"].width = 24
        for row in sheet.iter_rows():
            if row[0].value in {"Отчет по выбросам", "Сводка компании", "Итоговые показатели"}:
                row[0].font = Font(bold=True, size=13)
            elif row[0].value:
                row[0].font = Font(bold=True)
            if len(row) > 1:
                row[1].number_format = "#,##0.00"

    def _style_table(self, sheet, header_row):
        header_fill = PatternFill(fill_type="solid", fgColor="EAF4ED")
        for cell in sheet[header_row]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
        sheet.freeze_panes = sheet.cell(header_row + 1, 1)
        widths = [20, 24, 32, 14, 22, 16, 12, 14, 16, 12, 12, 12, 14, 12, 18, 22]
        for index, width in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(index)].width = width
        for row in sheet.iter_rows(min_row=header_row + 1):
            if row[0].value:
                row[0].number_format = "dd.mm.yyyy hh:mm"
            for cell, header in zip(row, TRIP_EXPORT_HEADERS, strict=True):
                if header in NUMERIC_FORMATS and isinstance(cell.value, int | float):
                    cell.number_format = NUMERIC_FORMATS[header]

    def _save(self, workbook):
        buffer = BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    def _period_label(self, filters):
        if filters.date_from and filters.date_to:
            return f"{filters.date_from:%d.%m.%Y} - {filters.date_to:%d.%m.%Y}"
        if filters.date_from:
            return f"с {filters.date_from:%d.%m.%Y}"
        if filters.date_to:
            return f"по {filters.date_to:%d.%m.%Y}"
        return "весь период"


def _summary_for_trips(rows):
    values = []
    for row in rows:
        item = {header: row[index].value for index, header in enumerate(TRIP_EXPORT_HEADERS)}
        values.append(item)
    if not values:
        return {}
    return {
        "Расстояние, км": _sum(values, "Расстояние, км"),
        "Время, мин": _sum(values, "Время, мин"),
        "Топливо, л": _sum(values, "Топливо, л"),
        "Стоимость, руб": _sum(values, "Стоимость, руб"),
        "CO2, кг": _sum(values, "CO2, кг"),
        "NOx, г": _sum(values, "NOx, г"),
        "PM, г": _sum(values, "PM, г"),
        "Эко-рейтинг": _avg(values, "Эко-рейтинг"),
        "CO2/км": _avg(values, "CO2/км"),
        "CO2/тонно-км": _avg(values, "CO2/тонно-км"),
    }


def _sum(values, key):
    return sum((value[key] for value in values if isinstance(value[key], int | float)), 0)


def _avg(values, key):
    numbers = [value[key] for value in values if isinstance(value[key], int | float)]
    if not numbers:
        return None
    return sum(numbers) / len(numbers)


def _cell_value(value):
    if isinstance(value, Decimal):
        return _number(value)
    if value is None:
        return MISSING_VALUE
    return value


def _number(value):
    if isinstance(value, Decimal):
        return float(value)
    return value


def _optional_number(value):
    if value is None:
        return MISSING_VALUE
    return _number(value)


def _excel_datetime(value):
    if not value:
        return MISSING_VALUE
    return timezone.localtime(value).replace(tzinfo=None)


def _trip_date(trip):
    return trip.actual_finish_at or trip.actual_start_at or trip.planned_start_at or trip.created_at


def _user_display_name(user):
    return user.get_full_name() or user.username


def report_summary_for_trips(trips):
    trips = list(trips)
    return {
        "trips_count": len(trips),
        "distance_km": sum((trip.route_option.distance_km for trip in trips), Decimal("0.00")),
        "fuel_liters": sum((trip.route_option.fuel_liters for trip in trips), Decimal("0.00")),
        "cost_rub": sum((trip.route_option.cost_rub for trip in trips), Decimal("0.00")),
        "co2_kg": sum((trip.route_option.co2_kg for trip in trips), Decimal("0.00")),
        "nox_g": sum((trip.route_option.nox_g for trip in trips), Decimal("0.00")),
        "pm_g": sum((trip.route_option.pm_g for trip in trips), Decimal("0.000")),
        "average_co2_kg_per_km": display_decimal(
            average_decimal((co2_kg_per_km(trip.route_option) for trip in trips), "0.001"),
            "0.001",
        ),
        "average_co2_kg_per_ton_km": display_decimal(
            average_decimal((co2_kg_per_ton_km(trip.route_option) for trip in trips), "0.0001"),
            "0.0001",
        ),
        "average_eco_rating": display_decimal(
            average_decimal((trip.route_option.eco_rating for trip in trips), "0.01"),
            "0.01",
        ),
        "toll_routes_count": sum(1 for trip in trips if has_tolls(trip.route_option)),
    }
